import openpyxl, os
from openpyxl import load_workbook

def template (path="", filename=""):
    #load excle file 
    wb = load_workbook(os.path.join(path, filename))
    ws = wb.worksheets[0]

    # add content
    ws['A1'] = 'FILENAME'
    ws['A2'] = 'KEYFIELD'
    ws['A3'] = 'KEYFIELD2'
    ws['A4'] = 'KEYFIELD3'
    ws['A5'] = 'KEYFIELD4'
    ws['A6'] = 'OVERWRITE'
    ws['B1'] = 'FUEL_ISSUES_FW'
    ws['B2'] = 'VEHICLE_ID_FW'
    ws['B3'] = 'TRANSACTION_DATE_FW'
    ws['B4'] = 'TRANSACTION_TIME_FW'
    ws['B5'] = 'AMOUNT_FW'
    ws['B6'] = 'NO'

    #save excel file
    wb.save(os.path.join(path, filename))

if __name__ =='__main__':
    template()